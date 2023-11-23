from django.utils import timezone
from django.shortcuts import render, redirect, get_object_or_404
from .models import dataSiswa, RekapPembayaran, pengaturan
from .forms import PendaftaranSiswaForm, RekapPembayaranForm
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q
import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse
from django.db.models import Sum
from datetime import datetime
# Create your views here.



def daftar_siswa(request):
    tahun_sekarang = timezone.now().year  # Ambil tahun sekarang
    print(tahun_sekarang)
    siswa = dataSiswa.objects.filter(tahun=tahun_sekarang).order_by('noPendaftaran')  # Mengurutkan siswa berdasarkan noPendaftaran secara ascending
    query = request.GET.get('q')
    
    if query:
        siswa = siswa.filter(
            Q(nama__icontains=query) |
            Q(noPendaftaran__icontains=query)|
            Q(nik__icontains=query)|
            Q(nisn__icontains=query)
        )
  

    paginator = Paginator(siswa, 10)
    
    page_number = request.GET.get('page')
    
    page = paginator.get_page(page_number)
    context = { 'page': page, 'query': query, 'tahun': tahun_sekarang  }
    return render(request, 'home.html', context)

@login_required
def edit_siswa(request, pk):
    siswa = get_object_or_404(dataSiswa, pk=pk)
    print(siswa.tahun)
    if request.method == 'POST':
        form = PendaftaranSiswaForm(request.POST, instance=siswa)
        if form.is_valid():
            # Ambil data yang sudah ada dalam database
            existing_siswa = dataSiswa.objects.filter(
                Q(noPendaftaran=form.cleaned_data['noPendaftaran']) |
                Q(nik=form.cleaned_data['nik']) |
                Q(nisn=form.cleaned_data['nisn'])
            ).exclude(pk=siswa.pk).first()
            
            if existing_siswa:
                return render(request, 'edit_siswa.html', {
                    'form': form,
                    'error_message': 'Data No Pendaftaran, NIK, atau NISN sudah terdaftar.'
                })

            form.save()
            return redirect('daftar-siswa')
    else:
        form = PendaftaranSiswaForm(instance=siswa)
        
    context = {'form': form}
    
    return render(request, 'edit_siswa.html', context)

@login_required
def pendaftaran_siswa(request):
    if request.method == 'POST':
        form = PendaftaranSiswaForm(request.POST)
        if form.is_valid():
            nik = form.cleaned_data.get('nik')
            nisn = form.cleaned_data.get('nisn')
            no_pendaftaran = form.cleaned_data.get('noPendaftaran')

            # Set nilai tahun secara otomatis
            tahun = datetime.now().year  # Menggunakan tahun saat ini

            # Periksa apakah data dengan NIK, NISN, atau No Pendaftaran yang sama sudah ada
            existing_nik = dataSiswa.objects.filter(nik=nik, tahun=tahun).exists()
            existing_nisn = dataSiswa.objects.filter(nisn=nisn, tahun=tahun).exists()
            existing_no_pendaftaran = dataSiswa.objects.filter(noPendaftaran=no_pendaftaran, tahun=tahun).exists()

            if existing_nik or existing_nisn or existing_no_pendaftaran:
                error_message = "NIK, NISN, atau No Pendaftaran sudah terdaftar."
                return render(request, 'pendaftaran.html', {'form': form, 'error_message': error_message})

            # Set tahun sebelum menyimpan data
            form.instance.tahun = tahun

            form.save()
            return redirect('daftar-siswa')
    else:
        form = PendaftaranSiswaForm()

    return render(request, 'pendaftaran.html', {'form': form})


def user_login(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username , password=password)
            if user is not None:
                login(request,user)
                return redirect('daftar-siswa')
    else:
        form =AuthenticationForm()
            
    return render(request, 'login.html' , {'form': form})
    
def user_logout(request):
    logout(request)
    return redirect('daftar-siswa')

@login_required
def delete_siswa(request , pk):
    siswa = get_object_or_404(dataSiswa, pk=pk)
    
    if request.method == 'POST':
        siswa.delete()
        return redirect('daftar-siswa')
    
    
    return render(request, 'delete.html', {'siswa': siswa})



def export_to_excel(request):
    tahun_sekarang = timezone.now().year
    siswa = dataSiswa.objects.filter(tahun =tahun_sekarang)

    # Buat workbook dan worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daftar Siswa"

    # Buat header kolom
    columns = [
        'No Pendaftaran',
        'Nama',
        'NISN',
        'NIK',
        'Asal Sekolah',
    ]

    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = Font(bold=True)

    # Isi data ke worksheet
    for siswa_data in siswa:
        row_num += 1
        row = [
            siswa_data.noPendaftaran,
            siswa_data.nama,
            siswa_data.nisn,
            siswa_data.nik,
            siswa_data.asal_sekolah,
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value

    # Konfigurasi response
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="daftar_siswa.xlsx"'

    # Simpan workbook ke response
    wb.save(response)

    return response
@login_required
def input_rekap_pembayaran(request, pk):
    siswa = get_object_or_404(dataSiswa, pk=pk)
    
    if request.method == 'POST':
        form = RekapPembayaranForm(request.POST)
        if form.is_valid():
            rekap_pembayaran = form.save(commit=False)
            rekap_pembayaran.siswa = siswa
            rekap_pembayaran.save()
            return redirect('detail-siswa', pk=siswa.pk)
    else:
        form = RekapPembayaranForm()

    context = {'form': form, 'siswa': siswa}
    return render(request, 'pembayaran.html', context)


@login_required
def detail_siswa(request, pk):
    siswa = get_object_or_404(dataSiswa, pk=pk)
    settings = get_object_or_404(pengaturan, pk=1)
    
    

    rekap_pembayaran = RekapPembayaran.objects.filter(siswa=siswa).order_by('-tanggal_rekap')
    total_pembayaran = sum(rekap.total_pembayaran for rekap in rekap_pembayaran)
    kekurangan_pembayaran = settings.total_pembayaran - total_pembayaran
    context = {
        'siswa': siswa,
        'rekap_pembayaran': rekap_pembayaran,
        'total_pembayaran': total_pembayaran,
        'kekurangan_pembayaran': kekurangan_pembayaran,
        'settings': settings
    }

    return render(request, 'detail_siswa.html', context)
@login_required
def rekap_pembayaran(request):
    tahun_sekarang = timezone.now().year
    # Query atau logika bisnis untuk mengambil data semua anak dan total pembayaran mereka
    rekap_pembayaran = dataSiswa.objects.annotate(total_pembayaran=Sum('rekappembayaran__total_pembayaran')).filter(tahun=tahun_sekarang).order_by('noPendaftaran')
    for rekap in rekap_pembayaran:
        print(rekap.noPendaftaran)

    context = {'rekap_pembayaran': rekap_pembayaran}
    return render(request, 'rekap_pembayaran.html', context)

def export_rekap_pembayaran_to_excel(request):
    tahun_sekarang = timezone.now().year
    rekap_pembayaran = dataSiswa.objects.annotate(total_pembayaran=Sum('rekappembayaran__total_pembayaran')).filter(tahun=tahun_sekarang).order_by('noPendaftaran')

    # Buat workbook dan worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rekap Pembayaran Siswa"

    # Buat header kolom
    columns = [
        'No Pendaftaran',
        'Nama',
        'Total Pembayaran',
    ]

    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = Font(bold=True)

    # Isi data ke worksheet
    for rekap_data in rekap_pembayaran:
        row_num += 1
        row = [
            rekap_data.noPendaftaran,
            rekap_data.nama,
            rekap_data.total_pembayaran,
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value

    # Konfigurasi response
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="rekap_pembayaran.xlsx"'

    # Simpan workbook ke response
    wb.save(response)

    return response